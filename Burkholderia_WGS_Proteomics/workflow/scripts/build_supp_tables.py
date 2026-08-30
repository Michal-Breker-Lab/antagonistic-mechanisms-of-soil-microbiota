#!/usr/bin/env python3
"""Build the supplementary tables: raw and core-centred LFQ, DE across all
contrasts, candidate toxins, functional enrichment and secretion systems.
"""
import csv
import math
import os
import re
import sys

import pandas as pd

from supp_xlsx_style import render_xlsx

sm = snakemake  # noqa: F821
sys.stderr = sys.stdout = open(sm.log[0], "w")

ROOT = os.path.abspath(sm.params.root)
OUT = os.path.dirname(sm.output[0])

LFQ = os.path.join(ROOT, "Proteomics/MF6_maxlfq_intensity.tsv")
LFQ_CENTRED = os.path.join(ROOT, "DE/day2_log2_maxlfq_core_centred.tsv")
CORE_SET = os.path.join(ROOT, "DE/day2_core_proteins.tsv")
OFFSETS = os.path.join(ROOT, "DE/day2_median_offsets.tsv")
GFF = os.path.abspath(sm.input.gff)
BIOLIB = os.path.abspath(sm.input.biolib)
CAND = os.path.abspath(sm.input.candidates)
PRED = os.path.abspath(sm.input.pred)
HEATMAP = os.path.join(ROOT,
                       "toxins/plots/candidates_sup_horizontal_abundance_heatmap.tsv")
ENRICH = os.path.join(ROOT, "DE/MF6/enrichment")
CLUSTER = os.path.join(ROOT, "Clustermap")
CLUSTER_FILES = [("COG", "core_cog_enrichment.tsv"),
                 ("KEGG", "core_pathway_enrichment.tsv")]
TXS_RUNS = os.path.abspath(sm.params.txsscan_runs)
TXS_SYS = os.path.abspath(sm.input.txsscan_systems)
TXS_STAGED = os.path.abspath(sm.input.txsscan_tree)
TXS_MF6 = os.path.abspath(sm.input.txss)
IMG_INFO = os.path.join(TXS_STAGED, "IMG_info.csv")
TXS_MAP = os.path.join(os.path.dirname(TXS_RUNS), "gembase")
ISOLATES = list(sm.params.isolates)
PLOT_VARIANT = sm.params.variant
PLOT_ALPHA = float(sm.params.alpha)
CONTRAST_RENAME = {"coculture_d2": "MF6_C+_D2_vs_MF6_C-_D2"}
UP_GROUP = CONTRAST_RENAME["coculture_d2"].split("_vs_")[0]
DIR_LABEL = {"up": f"Up in {UP_GROUP}", "down": f"Down in {UP_GROUP}"}


CONTRASTS = [
    ("MF6_withC_vs_alone", "DE/MF6/DE_coculture_d2.tsv",
     "DE/MF6/on_off_withC_d2_vs_alone_d2.tsv"),
    ("27D6_withC_vs_MF6_withC", "DE/27D6/DE_27D6_withC_vs_MF6_withC.tsv",
     "DE/27D6/on_off_27D6_withC_vs_MF6_withC.tsv"),
    ("27D6_alone_vs_MF6_alone", "DE/27D6/DE_27D6_alone_vs_MF6_alone.tsv",
     "DE/27D6/on_off_27D6_alone_vs_MF6_alone.tsv"),
    ("27D6_withC_vs_alone", "DE/27D6/DE_27D6_withC_vs_alone.tsv",
     "DE/27D6/on_off_27D6_withC_vs_alone.tsv"),
    ("34F7_withC_vs_MF6_withC", "DE/34F7/DE_34F7_withC_vs_MF6_withC.tsv",
     "DE/34F7/on_off_34F7_withC_vs_MF6_withC.tsv"),
    ("34F7_alone_vs_MF6_alone", "DE/34F7/DE_34F7_alone_vs_MF6_alone.tsv",
     "DE/34F7/on_off_34F7_alone_vs_MF6_alone.tsv"),
    ("34F7_withC_vs_alone", "DE/34F7/DE_34F7_withC_vs_alone.tsv",
     "DE/34F7/on_off_34F7_withC_vs_alone.tsv"),
]
LAST3 = ["DeepLocPro_localisation", "toxin_Pfam_domains",
         "T6SS_effector_probability"]
SCI_COLS = set()


def kegg_id(term):
    """One KEGG namespace for the whole `term` column.

    The DE enrichment reports pathways as `map00010`, the cluster enrichment as
    `ko00010`.  A KEGG pathway id is a prefix (the view: map / ko / ec / rn) plus a
    5-digit number that identifies the pathway, so those two are the SAME pathway -
    but as written they share no literal ids, and 111 pathways therefore appear
    twice in one column under two names, breaking any join or group-by on `term`.
    Normalised to `map` for both."""
    m = re.match(r"^[a-z]+(\d{4,6})$", term or "")
    return f"map{m.group(1)}" if m else term


SAMPLE_RE = re.compile(r"^(?P<strain>[^_]+)_(?P<cond>withC|alone)_d(?P<day>[23])"
                       r"_(?P<rep>[1-4])$")
GROUP_RE = re.compile(r"^((?P<strain>27D6|34F7|MF6)_)?(?P<cond>withC|alone)"
                      r"_d(?P<day>[23])$")


def sample_name(col):
    """`MF6_withC_d2_1` -> `MF6_C+_D2_1`; anything else unchanged."""
    m = SAMPLE_RE.match(col)
    if not m:
        return col
    return (f"{m['strain']}_{'C+' if m['cond'] == 'withC' else 'C-'}"
            f"_D{m['day']}_{m['rep']}")


def group_name(g):
    """`MF6_withC_d2` -> `MF6_C+_D2`.

    The wild-type ON/OFF table writes its groups WITHOUT the strain (`withC_d2`),
    the mutant tables write them with it, so the strain is filled in as MF6 when
    it is missing - otherwise the same group would appear under two names."""
    g = (g or "").strip()
    m = GROUP_RE.match(g)
    if not m:
        return g
    return (f"{m['strain'] or 'MF6'}_{'C+' if m['cond'] == 'withC' else 'C-'}"
            f"_D{m['day']}")


def strip_desc(desc, pid=""):
    """`AC1V0C_00005 ... DnaA` -> `... DnaA`.

    FragPipe repeats the locus tag inside the description. The trailing
    [organism] is already gone from the staged resource, and must NOT be stripped
    here: a trailing bracket is now part of the enzyme name, as in
    `superoxide dismutase [Cu-Zn]`."""
    if not desc:
        return ""
    d = desc.strip()
    if pid:
        d = re.sub(rf"\b{re.escape(pid)}\b", " ", d)
    return re.sub(r"\s{2,}", " ", d).strip()


def fix(v):
    """Kept as the single point where a cell could be rewritten; currently a no-op."""
    return v


def rd(path, delim="\t"):
    with open(os.path.join(ROOT, path) if not os.path.isabs(path) else path) as fh:
        return list(csv.DictReader((l for l in fh if not l.startswith("#")),
                                   delimiter=delim))


def chrom_label(seqid):
    """chr1 -> "chromosome 1", the form the published supplementary tables use.

    The pipeline works in the submission naming (chr1/chr2/chr3) everywhere else -
    it is what the GFF, the FASTA and config's absent_replicon say.  Only the reader-
    facing tables spell it out, so the mapping lives here, at the point the value is
    read for display, rather than being applied to the data itself."""
    n = seqid[3:] if seqid.startswith("chr") else ""
    return f"chromosome {n}" if n.isdigit() else seqid


def coords():
    """locus_tag -> (chromosome, start, stop, strand, product)."""
    out = {}
    with open(GFF) as fh:
        for line in fh:
            if line.startswith("#"):
                continue
            f = line.rstrip("\n").split("\t")
            if len(f) < 9 or f[2] != "CDS":
                continue
            a = dict(kv.split("=", 1) for kv in f[8].split(";") if "=" in kv)
            tag = a.get("locus_tag")
            if tag and tag not in out:
                out[tag] = (chrom_label(f[0]), int(f[3]), int(f[4]), f[6],
                            fix(a.get("product", "").replace("%2C", ",")))
    return out


def annotation():
    """The three columns S2 and S3 both end with, keyed by locus tag."""
    loc = {}
    for r in rd(BIOLIB):
        p = re.sub(r"^gnl\|extdb\|", "", r["protein_id"])
        loc[p] = r["deeplocpro_localization"]

    tox = {}
    for r in rd(CAND):
        acc = (r.get("tox_pfams") or "").strip()
        nam = (r.get("tox_pfam_names") or "").strip()
        if acc:
            tox[r["MF6_ID"]] = f"{acc} ({nam})" if nam else acc

    pred = {r["protein_id"]: round(float(r["xg_probas"]), 3)
            for r in rd(PRED) if r["protein_id"]}
    return loc, tox, pred


LFQ_DP = int(sm.params.lfq_decimals)


def dp4(v):
    """Fixed 4 decimals, so an LFQ column has ONE text form.

    Without this the column mixes representations: R drops trailing zeros, so a
    value that lands whole is written `22` among neighbours like `23.6256`, and
    Python's repr does the same for `24.0` vs `23.919`.  The source is rounded to
    4 dp already (max 4 decimals observed), so pinning the width is lossless."""
    if v in ("", None):
        return ""
    try:
        return f"{float(v):.{LFQ_DP}f}"
    except (TypeError, ValueError):
        return v


def log2(v):
    """MaxLFQ writes 0.0 for 'not quantified'; that is a blank, not a -inf."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return ""
    return dp4(math.log2(f)) if f > 0 else ""


def table_s1(co):
    rows = rd(LFQ)
    samples = [c for c in rows[0] if re.search(r"_d[23]_[1-4]$", c)]
    out = []
    for r in rows:
        c = co.get(r["Protein"], ("", "", "", "", ""))
        rec = {"Protein": r["Protein"],
               "Description": strip_desc(fix(r["Description"]), r["Protein"]),
               "chromosome": c[0], "start": c[1], "stop": c[2], "strand": c[3],
               "protein_length_aa": r["Protein Length"],
               "unique_spectra": r["Combined Unique Spectral Count"]}
        for col in samples:
            rec[sample_name(col)] = log2(r[col])
        out.append(rec)
    return out, [sample_name(c) for c in samples]


def description(xw, entries):
    """First sheet: what each table is, plus the conventions that span them.

    Row and column counts are read off the frames rather than typed in, so the
    sheet cannot drift out of date when a table changes."""
    rows = [{"item": "MF6 proteomics - supplementary tables", "title": "",
             "file": "", "size": "",
             "detail": "One sheet per table. The same tables are also on disk as "
                       "TSVs next to this workbook; see supp_tables/README.md for "
                       "the full method notes."},
            {"item": "", "title": "", "file": "", "size": "", "detail": ""}]
    for tag, title, fname, df, text in entries:
        rows.append({"item": f"Table {tag}", "title": title, "file": fname,
                     "size": f"{len(df):,} rows x {len(df.columns)} cols",
                     "detail": text})
    rows += [
        {"item": "", "title": "", "file": "", "size": "", "detail": ""},
        {"item": "Sample naming", "title": "", "file": "", "size": "",
         "detail": "<strain>_<C+|C->_D<day>_<replicate>. C+ = with "
                   "Chlamydomonas; C- = axenic. MF6_C+_D2_1 is wild type, "
                   "co-culture, day 2, replicate 1."},
        {"item": "Chromosomes", "title": "", "file": "", "size": "",
         "detail": "Numbered by length: chr1 (3.59 Mb), chr2 (2.98 Mb), "
                   "chr3 (1.18 Mb), as in the NCBI submission. The numbering does "
                   "not follow the contig numbering."},
        {"item": "Species", "title": "", "file": "", "size": "",
         "detail": "MF6 is Burkholderia sola. The PGAP annotation says "
                   "B. cenocepacia; that name has been corrected throughout "
                   "these tables."},
        {"item": "Blank cells", "title": "", "file": "", "size": "",
         "detail": "A blank is 'not measured', never zero: not quantified in that "
                   "run (S1, S1.1), detected in no replicate of that group (S3), "
                   "or a statistic the analysis does not report."},
        {"item": "Commas", "title": "", "file": "", "size": "",
         "detail": "Avoided so the TSVs load as CSV unquoted: list separators are "
                   "pipes and punctuation commas are semicolons. Commas inside "
                   "chemical names are kept - amylo-alpha-1;6-glucosidase would "
                   "be a different molecule."},
        {"item": "Contact / provenance", "title": "", "file": "", "size": "",
         "detail": "Built by scripts/supp_tables/build_supp_tables.py; verified by "
                   "scripts/supp_tables/check_types.py."},
    ]
    pd.DataFrame(rows, columns=["item", "title", "file", "size",
                                "detail"]).to_excel(
        xw, sheet_name="Description", index=False)
    ws = xw.sheets["Description"]
    for col, width in (("A", 22), ("B", 34), ("C", 42), ("D", 20), ("E", 108)):
        ws.column_dimensions[col].width = width
    from openpyxl.styles import Alignment
    for row in ws.iter_rows(min_col=5, max_col=5):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")


def table_s1_1(co):
    """The core-set-centred matrix the DE models actually ran on.

    Same shape of table as S1, different numbers and a different sample set:
    S1 is RAW log2 (comparable within a run, not across runs), this one is centred
    on each run's median over the core set (comparable across runs).

    It is day 2 only and 23 runs rather than 48, and that is not a subsetting
    choice made here: no centred matrix exists for day 3, because the DE analysis
    is day-2 only, and `34F7_alone_d2_1` is absent because the QC report fails it.
    The per-strain files under DE/<strain>/ are NOT merged in - each was
    centred within its own sample set, so their values are not on one scale."""
    rows = rd(LFQ_CENTRED)
    samples = [c for c in rows[0] if re.search(r"_d[23]_[1-4]$", c)]
    core = {r["Protein"] for r in rd(CORE_SET)}
    out = []
    for r in rows:
        c = co.get(r["Protein"], ("", "", "", "", ""))
        rec = {"Protein": r["Protein"],
               "Description": strip_desc(fix(r["Description"]), r["Protein"]),
               "chromosome": c[0], "start": c[1], "stop": c[2], "strand": c[3],
               "protein_length_aa": r["Protein Length"],
               "in_core_set": "yes" if r["Protein"] in core else "no"}
        for col in samples:
            rec[sample_name(col)] = dp4(r[col])
        out.append(rec)
    return out, [sample_name(c) for c in samples]


def table_s2(co, loc, tox, pred):
    """LONG format: one row per protein x contrast, only where there is a result.

    A wide table would need five columns per contrast and would still have to
    leave a cell empty for every protein a contrast never tested; long format
    says exactly which comparisons were made for which protein.  logFC is
    group_A - group_B."""
    out, unmapped = [], []
    for label, de_path, oo_path in CONTRASTS:
        de = {r["Protein"]: r for r in rd(de_path)}
        oo = ({r["Protein"]: r for r in rd(oo_path)}
              if os.path.exists(os.path.join(ROOT, oo_path)) else {})
        any_de = next(iter(de.values()), None)
        gA = group_name(any_de["group_A"]) if any_de else ""
        gB = group_name(any_de["group_B"]) if any_de else ""
        for p in sorted(set(de) | set(oo)):
            d, o = de.get(p), oo.get(p)
            c = co.get(p, ("", "", "", "", ""))
            src = d or o
            if d:
                nA, nB = d["n_valid_A"], d["n_valid_B"]
            else:
                pres = group_name(o["present_in"])
                if pres == gA:
                    nA, nB = o["n_valid_present"], o["n_valid_absent"]
                elif pres == gB:
                    nA, nB = o["n_valid_absent"], o["n_valid_present"]
                else:
                    nA = nB = ""
                    unmapped.append((p, label, pres))
            out.append({
                "Protein": p,
                "Description": strip_desc(fix(src.get("Description", "")), p) or c[4],
                "chromosome": c[0], "start": c[1], "stop": c[2], "strand": c[3],
                "contrast": f"{gA}_vs_{gB}" if gA else label,
                "group_A": gA, "group_B": gB,
                "logFC": d["logFC"] if d else "",
                "p_value": d["P.Value"] if d else "",
                "p_adj": d["adj.P.Val"] if d else "",
                "n_detected_A": nA,
                "n_detected_B": nB,
                "on_meanLFQ": (round(float(o["mean_log2_LFQ_present"]), 4)
                               if o else ""),
                LAST3[0]: loc.get(p, ""),
                LAST3[1]: tox.get(p, ""),
                LAST3[2]: pred.get(p, ""),
            })
    if unmapped:
        print(f"  WARNING: {len(unmapped)} ON/OFF rows whose present_in matched "
              f"neither group: {unmapped[:3]}")
    return out


MEAN_RE = re.compile(r"^(?P<strain>[^_]+)_Chl(?P<sign>[+-])_meanLog2LFQ$")


def mean_col(c):
    """`MF6_Chl-_meanLog2LFQ` -> `MF6_C-_D2_meanLog2LFQ`.

    Same naming as S1/S2 rather than a second dialect in the same supplement;
    the means are day 2 (source: DE/day2_log2_maxlfq_core_centred.tsv)."""
    m = MEAN_RE.match(c)
    return f"{m['strain']}_C{m['sign']}_D2_meanLog2LFQ" if m else c


def table_s3(co, loc, tox, pred):
    out, cols = [], []
    for r in rd(HEATMAP):
        p = r["protein_id"]
        c = co.get(p, ("", "", "", "", ""))
        rec = {"Protein": p, "Description": fix(r["label"]),
               "chromosome": c[0], "start": c[1], "stop": c[2], "strand": c[3],
               "in_pellet_matrix": r["in_pellet_matrix"]}
        for k, v in r.items():
            if MEAN_RE.match(k):
                rec[mean_col(k)] = dp4(v)
        rec[LAST3[0]] = loc.get(p, "")
        rec[LAST3[1]] = tox.get(p, "")
        rec[LAST3[2]] = pred.get(p, "")
        out.append(rec)
        cols = list(rec)
    return out, cols


def table_s4():
    """Two enrichment analyses in one table, distinguished by `analysis`.

    A  day-2 co-culture DE, the foreground the bubble figure plots
       (DE/MF6/enrichment/plots/bubble_coculture_d2.pdf): DE hits plus the
       strong/moderate ON/OFF proteins, i.e. the `de_plus_onoff` variant.  The
       other two variants give the same 13 significant terms and stay in the
       source TSVs.
    B  per-cluster enrichment of the core-set expression clusters
       (Clustermap/core_*_enrichment.tsv).

    They are NOT interchangeable and are labelled so nobody averages them: a
    cluster has no direction, its universe is the clustered and annotated proteins
    only (1,643 for COG, 907 for KEGG, against 3,549/1,549 for the DE test), and
    its BH correction runs within each cluster rather than within a contrast.
    """
    out = []
    cog_group = {}
    for ont, fname in CLUSTER_FILES:
        for r in rd(os.path.join(CLUSTER, fname)):
            if ont == "COG" and r["group"]:
                cog_group[r["category"]] = r["group"]

    for ont in ("COG", "KEGG"):
        for r in rd(os.path.join(ENRICH, f"{ont}_enrichment.tsv")):
            if r["variant"] != PLOT_VARIANT:
                continue
            out.append({
                "analysis": "day-2 co-culture DE",
                "set": CONTRAST_RENAME.get(r["contrast"], r["contrast"]),
                "direction": DIR_LABEL.get(r["direction"], r["direction"]),
                "ontology": r["ontology"], "term": kegg_id(r["term"]),
                "term_name": fix(r["term_name"]),
                "cog_group": cog_group.get(r["term"], "") if ont == "COG" else "",
                "k": r["k_fg"], "n": r["n_fg"], "K": r["K_bg"], "N": r["N_bg"],
                "fold_enrichment": r["fold_enrichment"],
                "p_value": r["p_value"], "p_adj": r["p_adj"],
                "proteins": r["proteins"].replace(",", "|"),
            })

    for ont, fname in CLUSTER_FILES:
        for r in rd(os.path.join(CLUSTER, fname)):
            out.append({
                "analysis": "core expression cluster",
                "set": f"cluster {r['cluster']}",
                "direction": "",
                "ontology": ont, "term": kegg_id(r["category"]),
                "term_name": fix(r["description"]),
                "cog_group": r["group"] if ont == "COG" else "",
                "k": r["k"], "n": r["n"], "K": r["K"], "N": r["N"],
                "fold_enrichment": round(float(r["fold_enrichment"]), 3),
                "p_value": r["pvalue"], "p_adj": r["FDR"],
                "proteins": r["protein_ids"].replace(";", "|"),
            })
    return out


def decomma(v):
    """Drop punctuation commas so the tables load as CSV unquoted.

    A comma FOLLOWED BY A SPACE is punctuation - `Replication, recombination and
    repair` - and becomes a semicolon.  A comma with no space after it belongs to a
    chemical name - `amylo-alpha-1,6-glucosidase`, `alpha,alpha-trehalase` - and is
    left alone, because deleting it silently renames the enzyme."""
    if not isinstance(v, str) or "," not in v:
        return v
    return re.sub(r",\s+", "; ", v)


def table_s5():
    """TXSScan hits, one row per protein per system.

    Relatives come from each genome's macsyfinder `best_solution.tsv`; MF6 comes
    from its own run, which is per replicon.  System-level columns are
    joined from all_systems.tsv, which already merges both runs, so wholeness and
    score mean the same thing on every row.

    A protein used by two systems gets a row per system - that is what
    macsyfinder reports and collapsing it would hide the sharing."""
    img = {}
    for r in rd(IMG_INFO, delim=","):
        k = {kk.strip(): (vv or "").strip() for kk, vv in r.items()}
        if k.get("Isolate ID"):
            img[k["Isolate ID"]] = k.get("IMG Genome ID", "")
    img["MF6"] = "This study"

    no_id = [g for g in ISOLATES if not img.get(g)]
    if no_id:
        sys.exit(f"no IMG Genome ID for {', '.join(no_id)}; add them to "
                 f"{os.path.basename(IMG_INFO)}")

    sysinfo = {}
    for r in rd(TXS_SYS):
        sysinfo[(r["genome"], r["system_id"])] = r

    out = []

    def add(genome, system_id, system, replicon, prot, component, status,
            ievalue, score, locus_bp="", gene_pos="", product="",
            msf_id="", gene_oid=""):
        si = sysinfo.get((genome, system_id), {})
        out.append({
            "Isolate ID": genome, "IMG Genome ID": img.get(genome, ""),
            "system": system, "system_id": system_id,
            "model_fqn": si.get("model_fqn", ""),
            "system_wholeness": si.get("wholeness", ""),
            "system_score": si.get("score", ""),
            "system_n_loci": si.get("n_loci", ""),
            "system_n_genes": si.get("n_genes", ""),
            "replicon": fix(replicon),
            "Protein": prot, "component": component, "hit_status": status,
            "i_evalue": ievalue, "hmm_score": score,
            "locus_bp": fix(locus_bp), "gene_position": gene_pos,
            "Description": product,
            "msf_protein_id": msf_id, "IMG_gene_oid": gene_oid,
        })

    for genome in ISOLATES:
        best = os.path.join(TXS_RUNS, genome, "best_solution.tsv")
        if not os.path.isfile(best):
            sys.exit(f"no best_solution.tsv for configured isolate {genome!r}: {best}")
        mp = os.path.join(TXS_MAP, f"{genome}.map.tsv")
        idmap = ({r["msf_id"]: r for r in rd(mp)} if os.path.isfile(mp) else {})
        for r in rd(best):
            m = idmap.get(r["hit_id"], {})
            add(genome, r["sys_id"], r["model_fqn"].rsplit("/", 1)[-1],
                r["replicon"], m.get("locus_tag") or r["hit_id"],
                r["gene_name"], r["hit_status"], r["hit_i_eval"], r["hit_score"],
                locus_bp=(f"{m['contig']}:{m['start']}-{m['end']}({m['strand']})"
                          if m.get("contig") else ""),
                gene_pos=r["hit_pos"],
                product=strip_desc(fix(m.get("product", ""))),
                msf_id=r["hit_id"], gene_oid=m.get("gene_oid", ""))

    for r in rd(TXS_MF6):
        add("MF6", r["system_id"], r["system"], r["replicon"], r["MF6_ID"],
            r["component"], r["hit_status"], r["i_evalue"], r["hmm_score"],
            locus_bp=r["locus"], product=strip_desc(fix(r["PGAP_product"])),
            msf_id=r["MF6_ID"])
    return out


_DEST = dict(sm.params.render)
RENDER = {
    "Table_02_LFQ_raw_proteomics.tsv": (
        _DEST["Table_02_LFQ_raw_proteomics.tsv"], "log2 LFQ",
        "Raw log2 MaxLFQ intensities per protein per proteomic run for Bso MF6, "
        "27D6, and 34F7, Related to Figure 2"),
    "Table_03_LFQ_normalized_proteomics.tsv": (
        _DEST["Table_03_LFQ_normalized_proteomics.tsv"], "core-centred LFQ",
        "Core-set-centered log2 MaxLFQ intensities per protein, day 2 proteomic "
        "runs, Related to Figure 2"),
    "Table_01_secretion_systems.tsv": (
        _DEST["Table_01_secretion_systems.tsv"], "txsscan analysis",
        "Secretion system, pilus and flagellum components detected by TXSScan in "
        "Bso MF6 and its relatives, Related to Figure 2"),
    "Table_08_potential_toxins.tsv": (
        _DEST["Table_08_potential_toxins.tsv"], "Potential Toxins",
        "Candidate toxin and secretion-associated proteins in the Bso MF6 pellet "
        "proteome, Related to Figure 3"),
}

def write(name, rows, cols=None):
    path = os.path.join(OUT, name)
    cols = cols or list(rows[0])
    rows = [{k: (v if k in SCI_COLS else decomma(v)) for k, v in r.items()}
            for r in rows]
    with open(path, "w", newline="") as fh:
        w = csv.DictWriter(fh, cols, delimiter="\t", lineterminator="\n",
                           extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  {name}: {len(rows)} rows x {len(cols)} cols")
    if name in RENDER:
        xlsx, sheet, title = RENDER[name]
        render_xlsx(path, xlsx, sheet, title)
    return typed_frame(rows, cols)


def typed_frame(rows, cols):
    """One storage type per column for the .xlsx.

    Values reach here as a mix of pass-through strings from the source files and
    Python numbers from anything that was rounded, and pandas keeps that mix -
    which lands in Excel as some cells text and some numeric IN THE SAME COLUMN
    (fold_enrichment was 205 text + 971 numeric), so sorting and filtering that
    column in Excel silently splits it in two.  Each column is therefore coerced
    once: integer if every non-empty value is one, else float if they all parse,
    else text.  The TSV is untouched - it is text by nature and keeps the source
    precision."""
    df = pd.DataFrame(rows, columns=cols)
    for c in cols:
        vals = [v for v in df[c] if v not in ("", None) and v == v]
        if not vals:
            df[c] = df[c].astype("string")
            continue
        strs = [str(v) for v in vals]
        if all(re.fullmatch(r"-?\d+", x) for x in strs):
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
        else:
            try:
                [float(x) for x in strs]
                df[c] = pd.to_numeric(df[c], errors="coerce")
            except ValueError:
                df[c] = df[c].astype("string")
    return df


def main():
    os.makedirs(OUT, exist_ok=True)
    co = coords()
    loc, tox, pred = annotation()
    print(f"coordinates for {len(co)} CDS; DeepLocPro {len(loc)}; "
          f"toxin-Pfam proteins {len(tox)}; T6SS predictions {len(pred)}")

    s1, samples = table_s1(co)
    d1 = write("Table_02_LFQ_raw_proteomics.tsv", s1,
               ["Protein", "Description", "chromosome", "start", "stop",
                "strand", "protein_length_aa", "unique_spectra"] + samples)

    s11, samples11 = table_s1_1(co)
    d11 = write("Table_03_LFQ_normalized_proteomics.tsv", s11,
                ["Protein", "Description", "chromosome", "start", "stop",
                 "strand", "protein_length_aa", "in_core_set"] + samples11)
    print(f"    {sum(1 for r in s11 if r['in_core_set'] == 'yes')} of "
          f"{len(s11)} proteins are in the core set that defined the offsets")

    s2 = table_s2(co, loc, tox, pred)
    d2 = write("DE_all_contrasts.tsv", s2,
               ["Protein", "Description", "chromosome", "start", "stop", "strand",
                "contrast", "group_A", "group_B", "logFC", "p_value", "p_adj",
                "n_detected_A", "n_detected_B", "on_meanLFQ"] + LAST3)
    print(f"    {len(set(r['Protein'] for r in s2))} distinct proteins over "
          f"{len(set(r['contrast'] for r in s2))} contrasts")

    s3, cols3 = table_s3(co, loc, tox, pred)
    d3 = write("Table_08_potential_toxins.tsv", s3, cols3)

    s4 = table_s4()
    d4 = write("functional_enrichment.tsv", s4)
    from collections import Counter
    for a in sorted({r["analysis"] for r in s4}):
        rows_a = [r for r in s4 if r["analysis"] == a]
        sig = [r for r in rows_a if float(r["p_adj"]) < PLOT_ALPHA]
        print(f"    {a}: {len(rows_a)} terms, {len(sig)} at BH<{PLOT_ALPHA} "
              f"({dict(Counter(r['ontology'] for r in sig))})")

    s5 = table_s5()
    d5 = write("Table_01_secretion_systems.tsv", s5)

    xlsx = os.path.join(OUT, "MF6_supplementary_tables.xlsx")
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        description(xw, [
            ("S1", "Raw log2 LFQ", "Table_02_LFQ_raw_proteomics.tsv", d1,
             "log2 RAW MaxLFQ per protein per run; all 48 runs. Comparable "
             "within a run but NOT across runs - use S1.1 for that. Blank = not "
             "quantified in that run. Includes the two QC-failed runs "
             "(34F7_C-_D2_1 and 34F7_C-_D3_1) which every analysis excludes."),
            ("S1.1", "Core-centred log2 LFQ",
             "Table_03_LFQ_normalized_proteomics.tsv", d11,
             "The same values centred on each run's median over the core set - "
             "the matrix the DE models ran on; comparable across runs. Day 2 "
             "only and 23 runs: no centred matrix exists for day 3 and the "
             "QC-failed 34F7_C-_D2_1 is out. in_core_set marks the proteins that "
             "defined the offsets (DE/day2_median_offsets.tsv)."),
            ("S2", "Differential expression and ON/OFF",
             "DE_all_contrasts.tsv", d2,
             "LONG format - one row per protein per contrast, seven day-2 "
             "contrasts. logFC is group_A minus group_B. on_meanLFQ is the ON/OFF "
             "column: proteins present in one group and never detected in the "
             "other carry no logFC because limma cannot fit them, so this is the "
             "only place they appear."),
            ("S3", "Potential toxins", "Table_08_potential_toxins.tsv", d3,
             "Mean core-centred log2 LFQ per strain x condition for the 36 "
             "candidate proteins behind the abundance heatmap. Blank = detected "
             "in no replicate of that group. Chromosome 3 is censored in the 27D6 "
             "and 34F7 columns - both mutants lost that replicon."),
            ("S4", "Functional enrichment",
             "functional_enrichment.tsv", d4,
             "Two enrichment analyses, told apart by `analysis`: the day-2 "
             "co-culture DE contrast and the eight core expression clusters. "
             "Different universes and separate BH corrections, so a p_adj from "
             "one is not comparable with a p_adj from the other."),
            ("S5", "Secretion systems", "Table_01_secretion_systems.tsv", d5,
             "TXSScan/MacSyFinder hits, one row per protein per system, for MF6 "
             f"and {len(ISOLATES)} relatives."),
        ])
        d1.to_excel(xw, sheet_name="S1 log2 LFQ", index=False)
        d11.to_excel(xw, sheet_name="S1.1 core-centred LFQ", index=False)
        d2.to_excel(xw, sheet_name="S2 DE and ON-OFF", index=False)
        d3.to_excel(xw, sheet_name="S3 Potential Toxins", index=False)
        d4.to_excel(xw, sheet_name="S4 enrichment", index=False)
        d5.to_excel(xw, sheet_name="S5 secretion systems", index=False)
    import openpyxl
    names = openpyxl.load_workbook(xlsx, read_only=True).sheetnames
    print(f"  MF6_supplementary_tables.xlsx: {len(names)} sheets -> "
          + " | ".join(names))

    left = sum(1 for f in os.listdir(OUT) if f.endswith(".tsv")
               for line in open(os.path.join(OUT, f))
               if "cenocepacia" in line or re.search(r"\bMF6_\d{6}\b|\bcontig_[124]\b", line))
    print(f"\nlines still carrying an old id / contig / species: {left}")
    assert left == 0, "the staged resources should already be relabelled"
    for f in sorted(x for x in os.listdir(OUT) if x.endswith(".tsv")):
        txt = open(os.path.join(OUT, f)).read()
        punct = len(re.findall(r",\s", txt))
        chem = len(re.findall(r",\S", txt))
        print(f"  {f}: {punct} punctuation commas, {chem} inside chemical names")
    return 0


if __name__ == "__main__":
    sys.exit(main())
