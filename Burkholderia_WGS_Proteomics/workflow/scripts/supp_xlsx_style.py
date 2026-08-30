#!/usr/bin/env python3
"""Shared styling for the supplementary workbooks."""
import os

from openpyxl.styles import Alignment, Font

TITLE_FONT = Font(name="Arial", size=12, bold=True)
HEADER_FONT = Font(name="Calibri", size=12, bold=True)
BODY_FONT = Font(name="Calibri", size=12)
HEADER_ALIGN = Alignment(vertical="top", wrap_text=True)


def write_header(ws, title, columns, title_row=1, header_row=3):
    """Title in A1, blank row, column names on row 3 - the published layout."""
    c = ws.cell(title_row, 1)
    c.value = title
    c.font = TITLE_FONT
    for j, name in enumerate(columns, start=1):
        h = ws.cell(header_row, j)
        h.value = name
        h.font = HEADER_FONT
        h.alignment = HEADER_ALIGN


def style_body(ws, first_row=4):
    """Calibri 12 on every data cell.

    Applied after the rows are appended rather than per-cell as they are written:
    openpyxl creates a style object per assignment, and on a 22,811 x 24 sheet that
    is half a million of them.  One shared Font instance costs nothing.
    """
    for row in ws.iter_rows(min_row=first_row):
        for cell in row:
            cell.font = BODY_FONT


def render_xlsx(tsv_path, xlsx_path, sheet, title):
    """Render a computed TSV as one of the published workbooks.

    Title in A1, blank row, header on row 3, body below, panes frozen at A4 -
    the layout every supplementary table in this project uses.
    """
    import csv
    import openpyxl

    with open(tsv_path) as fh:
        rows = list(csv.reader((l for l in fh if not l.startswith("#")),
                               delimiter="\t"))
    if not rows:
        raise SystemExit(f"no rows in {tsv_path}")
    header, body = rows[0], rows[1:]

    def as_number(v):
        try:
            f = float(v)
        except ValueError:
            return None
        return int(f) if f.is_integer() and "." not in v and "e" not in v.lower() else f

    n_cols = len(header)
    numeric = []
    for j in range(n_cols):
        vals = [r[j] for r in body if j < len(r) and r[j] != ""]
        numeric.append(bool(vals) and all(as_number(v) is not None for v in vals))

    def cast(v, j):
        if v == "":
            return None
        return as_number(v) if numeric[j] else v

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    write_header(ws, title, header)
    for r in body:
        ws.append([cast(v, j) for j, v in enumerate(r)])
    style_body(ws)
    ws.freeze_panes = "A4"
    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb.save(xlsx_path)
    print(f"  {os.path.basename(xlsx_path)}: {len(body)} rows x {len(header)} cols")
